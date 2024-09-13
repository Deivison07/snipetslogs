from openpyxl import load_workbook

# Abre o arquivo Excel existente
caminho_arquivo = "dados_exemplo.xlsx"
workbook = load_workbook(caminho_arquivo)

# Seleciona a planilha desejada (substitua "NomeDaPlanilha" pelo nome da sua planilha)
sheet = workbook["Dados de Exemplo"]

# Dados a serem inseridos (você pode personalizar conforme necessário)
dados = [
    ["Nome", "Idade", "Cidade"],
    ["Alice", 30, "São Paulo"],
    ["Bob", 25, "Rio de Janeiro"],
    ["Carlos", 35, "Belo Horizonte"]
]

# Encontra a próxima linha vazia para inserir os dados
linha_inicial = sheet.max_row + 1  # A próxima linha vazia é após a última linha preenchida

# Insere os dados na planilha a partir da próxima linha vazia
for i, linha in enumerate(dados, start=linha_inicial):
    for j, valor in enumerate(linha, start=1):  # `start=1` porque a contagem de colunas começa em 1 no Excel
        sheet.cell(row=i, column=j, value=valor)

# Salva as alterações no arquivo Excel
workbook.save(caminho_arquivo)

print(f"Dados inseridos com sucesso na planilha '{caminho_arquivo}'.")
